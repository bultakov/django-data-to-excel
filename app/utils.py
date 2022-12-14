from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.writer.excel import save_virtual_workbook


def data_to_xlsx(user_data: list):
    wb = Workbook()
    ws = wb.active

    ws.auto_filter.ref = f"A1:D{len(user_data) + 1}"  # filter
    names: dict = {"A1": "ID", "B1": "Name", "C1": "Phone", "D1": "About"}
    for key, value in names.items():
        cl = ws[key]
        cl.value = value
        cl.alignment = Alignment(horizontal="center", vertical="center")
        cl.font = Font(b=True)

    for i, user in enumerate(user_data, start=2):
        for j, data in enumerate(user, start=1):
            cr = ws.cell(row=i, column=j)
            cr.value = data
    response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/xlsx')
    response['Content-Disposition'] = 'attachment; filename=user_data.xlsx'
    return response
